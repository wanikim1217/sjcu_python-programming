# https://openpyxl.readthedocs.io/en/stable/tutorial.html
# pip install openpyxl

# 용어 정리 : 워크북(엑셀파일) - 워크시트(시트) - 셀(행,열), Range : 셀의 범위

# 새로우 엑셀파일 맏들기
from openpyxl import Workbook, load_workbook
import random

wb = Workbook() # 새 워크북 생성
ws = wb.active # 현재 활성화된 sheet 가져옴
print(ws)

ws = wb.create_sheet("New-Sheet") # 새로운 sheet 생성
print(ws)

ws_list = wb.sheetnames # 모든 sheet 이름 확인
print(ws_list)

# 새로운 Cell 데이터 입력
ws["A1"] = "kor"
ws["B1"] = "math"
ws["C1"] = "grade"

for i in range(2,12):
  kor = random.randint(10, 100)
  math = random.randint(10, 100)
  ws.cell(row=i, column=1).value = kor
  ws.cell(row=i, column=2).value = math
  ws["C"+str(i)].value = "=IF(AVERAGE(A"+str(i)+":B"+str(i)+")>=70, \"A\", IF(AVERAGE(A"+str(i)+":B"+str(i)+")>=40, \"B\", \"C\"))"
  
wb.save(r"grade.xlsx") #결과 엑셀파일 저장

# 엑셀파일 불러오기
wb = load_workbook("grade.xlsx")
ws_list = wb.sheetnames
print(ws_list)

ws = wb["New-Sheet"]
print(ws)

range = ws["A1:C11"] # 모든 셀 가져오기
print(range)

for cell in range:
  print(cell[0].value, cell[1].value, cell[2].value)
